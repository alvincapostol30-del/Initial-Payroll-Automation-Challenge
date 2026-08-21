from openpyxl import Workbook
import csv
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


workbook = Workbook()

worksheet = workbook.active


#initialize header list
headers = [
    "Employee ID",
    "Employee Name",
    "Department",
    "Salary",
    "Status",
    "Tax",
    "Net Salary",
    "Bonus Rate",
    "Bonus Amount"
]

#Function block

def calculate_status(salary):
    if salary > 50000:
        return "Senior"
    else:
        return "Junior"

def calculate_tax(salary):
    if salary > 50000:
        return salary * 0.12
    else:
        return salary * 0.08

def calculate_net_salary(salary,tax):
    return salary - tax

def calculate_bonus_rate(salary):
    if salary > 50000:
        return "15%"
    else:
        return "10%"
def calculate_bonus_amount(salary):
    if salary >50000:
        return salary * 0.15
    else:
        return salary * 0.10



#get last index

last_header_list_index = len(headers)

#Write all headers
for r in range(1,last_header_list_index+1):
    worksheet.cell(1,r).value = headers[r-1]
    worksheet.cell(1,r).font = Font(bold = True,color="FFFFFFFF")
    worksheet.cell(1,r).fill = PatternFill(
    fill_type = "solid",
    start_color = "FF1F4E78")
#read csv file
with open("Input/employees.csv", newline="") as file:
    reader = csv.DictReader(file)

    #initialize first row counter value
    counter = 2
    for row in reader:
        worksheet.cell(counter,1).value = row["Employee ID"]
        worksheet.cell(counter,2).value = row["Employee Name"]
        worksheet.cell(counter,3).value = row["Department"]
        #convert salary from str to number
        salary = int(row['Salary'])  
        
        worksheet.cell(counter,4).value = salary
        worksheet.cell(counter,4).number_format = '₱#,##0.00'

        status = calculate_status(salary)
        tax = calculate_tax(salary)
        salary_net = calculate_net_salary(salary,tax)
        bonus_rate = calculate_bonus_rate(salary)
        bonus_amount = calculate_bonus_amount(salary)

        #write values to excel file 
        worksheet.cell(counter,5).value = status
        worksheet.cell(counter,6).value = tax
        worksheet.cell(counter,6).number_format = '₱#,##0.00'
        worksheet.cell(counter,7).value = salary_net
        worksheet.cell(counter,7).number_format = '₱#,##0.00'
        worksheet.cell(counter,8).value = bonus_rate
        worksheet.cell(counter,9).value = bonus_amount
        worksheet.cell(counter,9).number_format = '₱#,##0.00'

       

        counter+=1


#Create timestamp for filename
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f"Output/Payroll_{timestamp}.xlsx"

#auto fit columns
for column in worksheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)

    for cell in column:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))

    extra_width = 2

    if column_letter in ["D", "F", "G", "I"]:
        extra_width = 6

    worksheet.column_dimensions[column_letter].width = max_length + extra_width

worksheet.freeze_panes = "A2"

last_row = worksheet.max_row

table = Table(
    displayName="PayrollTable",
    ref=f"A1:I{last_row}"
)

style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

table.tableStyleInfo = style

worksheet.add_table(table)

#save file 
workbook.save(filename)


print(f"Payroll successfully generated: {filename}")

        
